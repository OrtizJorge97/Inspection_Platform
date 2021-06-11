from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import json
import sys
import os

from ..System.SystemService import SystemServices
from Constants.PartNumberConstants import PartNumberConstantsClass
from Helpers.PartNumberHelper.PartNumberModule import PartNumberClass, ConvertShowableInspectionIntoObject
from DeltaHubRestApi.Helpers.AddPartNumber.AddPartNumberHelper import *

class InspectionsSheetExcelService:
    def __init__(self, fileName):
        self.jsonInspectionSheet = None
        #self.fileName = None
        self.fileName = fileName
        self.filePath = None
        self.inspectionSheetPandas = None
        self.dataToShowInPlatform = None
        #private attribute
        self.__basicInformationCellCoords = {
            "Organization": [1, 2],
            "SupplierVendorCode": [2, 2],
            "PartNumber": [1, 9],
            "PartName": [2, 9],
            "DesignRecordChangeLevel": [3, 10],
            "EngineeringChangeDocuments": [4, 10],
            "INSPECTIONBY": [4, 0],
            "RawMaterial": [6, 1],
            "ApprovedBy": [6, 5],
            "Supplier": [6, 8],
            "Date": [6, 12]
        }
        self.__basicInformationData = {
            "Organization": "",
            "SupplierVendorCode": "",
            "PartNumber": "",
            "PartName": "",
            "DesignRecordChangeLevel": "",
            "EngineeringChangeDocuments": "",
            "INSPECTIONBY": "",
            "RawMaterial": "",
            "ApprovedBy": "",
            "Supplier": "",
            "Date": ""
        }
        self.basicInfoPandas = None
        self.purchaseOrder = ""
        self.dataToFillSection = None
        self.dataNotesSection = None
        self.dataToFillByAppSection = None
        self.errorMessage = None
        self.__itemNumberStartNotes = 1.0
        self.__itemNumberEndNotes = 2.0
        self.__defaultDataToFillNanValues = {
            "Item": " ",
            "Dimension": " ",
            "Units": " ",
            "SuperiorLimit": " ",
            "InferiorLimit": " ",
            "TestDate": " ",
            "QuantityTested": " ",
            "Measure1": " ",
            "Measure2": " ",
            "Measure3": " ",
            "Measure4": " ",
            "Measure5": " ",
            "Measure6": " ",
            "Ok": " ",
            "NotOk": ""
        }

    def AddFileExtensionIfNecessary(self):
        if ".xlsx" not in self.fileName and ".XLSX" not in self.fileName:
            self.filePath = f"{self.filePath}.xlsx"
            self.fileName = f"{self.fileName}.xlsx"
        """
        if ".XLSX" in self.filePath:
            self.filePath = self.filePath.replace(".XLSX", ".xlsx")
        """

    @staticmethod
    def TakeOutFileExtensionIfNecessary(purchaseOrderName):
        filename, file_extension = os.path.splitext(purchaseOrderName)
        if file_extension:
            purchaseOrderName = purchaseOrderName.replace(file_extension, "")

        return purchaseOrderName


    def ReadInspectionSheetExcel(self):
        systemServices = SystemServices(PartNumberConstantsClass.FOLDER_SAVE) #Buils only until path to dekstop

        self.filePath = systemServices.BuildPathForFile(self.fileName) #Build the path until fileName
        self.AddFileExtensionIfNecessary()
        print(self.filePath)

        if not systemServices.CheckIfFilePathExists(self.filePath):
            self.errorMessage = "File does not exist"
            self.jsonInspectionSheet = {
                "errorMessage": self.errorMessage
            }
            return
        print(f"File Path: {self.filePath}")
        self.inspectionSheetPandas = pd.read_excel(io=self.filePath)

    def ConvertInspectionSheetToJson(self):
        dataToFillStartIndex = 0
        floatSample = 1.3
        inspectionSheetPandasRows = self.inspectionSheetPandas.shape[0]
        inspectionSheetPandasColumns = self.inspectionSheetPandas.shape[1]

        print(self.inspectionSheetPandas)

        for row in list(range(inspectionSheetPandasRows)):
            dataToFillStartIndex = row
            value = self.inspectionSheetPandas.iloc[row, 0]

            if value is not None:
                isValid, itemValue = PartNumberClass.TryConvertToFloat(value)
                if isValid:
                    print(f"Row: {row} | Item: {itemValue} | isValid: {isValid}")
                    if itemValue >= self.__itemNumberStartNotes:
                        break
            """
            if type(self.inspectionSheetPandas.iloc[row, 0]) == type(floatSample):
                #print(f"Item: {self.inspectionSheetPandas.iloc[dataToFillStartIndex:, 0]}")
                if self.inspectionSheetPandas.iloc[row, 0] > self.__itemNumberStartNotes:
                    break
            """

        print(f"Index where numbers start!: {dataToFillStartIndex} | value: {self.inspectionSheetPandas.iloc[dataToFillStartIndex, 0] }")
        self.dataToFillSection = self.inspectionSheetPandas.iloc[dataToFillStartIndex:, 0:].copy()
        print(self.dataToFillSection)

        dataToFillRows = self.dataToFillSection.shape[0]
        dataToFillColumns = self.dataToFillSection.shape[1]

        #Analyze all the pandas section to fill
        notesIndexList = []
        for row in list(range(dataToFillRows)):
            item = self.dataToFillSection.iloc[row, 0]
            #print(self.dataToFillSection.iloc[row, 3])
            if float(item) >= self.__itemNumberStartNotes and float(item) < self.__itemNumberEndNotes:
                notesIndexList.append(row)
                print(float(item))


        #print(notesIndexList)
        #print(self.dataToFillSection)
        #print(len(notesIndexList))
        self.basicInfoPandas = PartNumberClass.ConvertBasicInformationIntoPandas(self.inspectionSheetPandas, self.__basicInformationData, self.__basicInformationCellCoords)

        #print(notesIndexList)

        self.dataNotesSection = self.dataToFillSection.iloc[notesIndexList[0]:len(notesIndexList), :]
        self.dataToFillByAppSection = self.dataToFillSection.iloc[len(notesIndexList):, :]
        self.dataToFillByAppSection = PartNumberClass.InitializeDataToFillByAppAsPandas(self.dataToFillByAppSection, self.__defaultDataToFillNanValues) #Initialize pandas with columns names and not with unamed


        #print("Basic Data Pandas")
        #print(self.basicInfoPandas)
        #print("data to fill by app section")
        #print(self.dataToFillByAppSection)

        """
        print(f"Part Number: \n{self.basicInfoPandas['PartNumber'].to_string(index=False).replace(' ', '')} ")
        """

        inspectionForm = {
            "basicInformation": {
                "Organization": self.basicInfoPandas["Organization"].to_string(index=False).replace(" ", ""),
                "SupplierVendorCode": self.basicInfoPandas["SupplierVendorCode"].to_string(index=False).replace(" ", ""),
                "PartNumber": self.basicInfoPandas["PartNumber"].to_string(index=False).replace(" ", ""),
                "PartName": self.basicInfoPandas["PartName"].to_string(index=False).replace(" ", ""),
                "PurchaseOrder": self.purchaseOrder,
                "DesignRecordChangeLevel": self.basicInfoPandas["DesignRecordChangeLevel"].to_string(index=False).replace(" ", ""),
                "EngineeringChangeDocuments": self.basicInfoPandas["EngineeringChangeDocuments"].to_string(index=False).replace(" ", ""),
                "RawMaterial": self.basicInfoPandas["RawMaterial"].to_string(index=False).replace(" ", ""),
                "ApprovedBy": self.basicInfoPandas["ApprovedBy"].to_string(index=False).replace(" ", ""),
                "Supplier": self.basicInfoPandas["Supplier"].to_string(index=False).replace(" ", ""),
                "Date": self.basicInfoPandas["Date"].to_string(index=False).replace(" ", "")
            },
            "notes": list(self.dataNotesSection.iloc[:, 1].fillna("")),
            "dataToFillByApp": {
                "item": list(self.dataToFillByAppSection["Item"]),
                "dimension": list(self.dataToFillByAppSection["Dimension"]), #string
                "units": list(self.dataToFillByAppSection["Units"]), #string
                "inferiorLimit": list(self.dataToFillByAppSection["InferiorLimit"]),
                "superiorLimit": list(self.dataToFillByAppSection["SuperiorLimit"]),
                "testDate": list(self.dataToFillByAppSection["TestDate"]),
                "quantityTested": list(self.dataToFillByAppSection["QuantityTested"]), #string
                "measure1": list(self.dataToFillByAppSection["Measure1"]),
                "measure2": list(self.dataToFillByAppSection["Measure2"]),
                "measure3": list(self.dataToFillByAppSection["Measure3"]),
                "measure4": list(self.dataToFillByAppSection["Measure4"]),
                "measure5": list(self.dataToFillByAppSection["Measure5"]),
                "measure6": list(self.dataToFillByAppSection["Measure6"]),
                "ok": list(self.dataToFillByAppSection["Ok"]),
                "notOk": list(self.dataToFillByAppSection["NotOk"])
            },
            "errorMessage": self.errorMessage
        }
        print(inspectionForm["dataToFillByApp"])
        self.dataToShowInPlatform = inspectionForm["dataToFillByApp"]

        self.jsonInspectionSheet = json.dumps(inspectionForm, indent=4)
        #print(inspectionForm["dataToFillByApp"]["item"])


class InspectionSheetOpenPyManager:
    #Second class being a patch
    def __init__(self):
        self.workbook = None
        self.worksheet = None
        self.mainDataFrame = None
        self.mainDataFrameCopy = None
        self.filePath = None
        self.itemColumnIndex = 0 #in pandas the column which item is, is 0
        self.itemWhereStarts = 2 #in inspections writable area starts where item number is 2 in document
        self.detectedWritableRow = None

        self.writableArea = None

    def ReadInspectionSheet(self, filePath):
        self.filePath = filePath
        self.workbook = load_workbook(filename=self.filePath)
        self.worksheet = self.workbook.active

    def UpdateInspectionSheet(self, postedDict):
        data = self.worksheet.values
        cols = next(data)[0:]

        self.mainDataFrame = PreprocessToMainDataFrame(data, cols)
        self.mainDataFrameCopy = self.mainDataFrame.copy()
        #print(self.mainDataFrameCopy)

        self.DetectWritableArea()
        self.WriteToDataFrame(postedDict)
        self.WriteToInspection()

    def DetectWritableArea(self):
        #print(self.mainDataFrameCopy.iloc[9:, self.itemColumnIndex])
        index = 0
        for itemValue in self.mainDataFrameCopy.iloc[9:, self.itemColumnIndex]:
            itemNumber = float(itemValue)
            if itemNumber >= self.itemWhereStarts:
                self.detectedWritableRow = index
                break
            index += 1

        self.writableArea = self.mainDataFrameCopy.iloc[(9 + self.detectedWritableRow):,:]
        #print(self.writableArea)

    def WriteToDataFrame(self, postedDict):

        self.writableArea.iloc[0:,7] = postedDict["Measure1"]
        self.writableArea.iloc[0:,8] = postedDict["Measure2"]
        self.writableArea.iloc[0:,9] = postedDict["Measure3"]
        self.writableArea.iloc[0:,10] = postedDict["Measure4"]
        self.writableArea.iloc[0:,11] = postedDict["Measure5"]
        self.writableArea.iloc[0:,12] = postedDict["Measure6"]

        self.writableArea.iloc[0:, 13] = postedDict["Ok"]
        self.writableArea.iloc[0:, 14] = postedDict["NotOk"]

        #print(self.mainDataFrameCopy)

    def WriteToInspection(self):
        rows = dataframe_to_rows(self.mainDataFrameCopy, index=False, header=False)
        # print(enumerate(rows, 1))
        # print(list(dataframe_to_rows(mainDataFrame)))
        # for r_idx, row in enumerate(rows, 1):
        #    print(f"row index: {r_idx} | value: {row}")

        for r_idx, row in enumerate(rows, 1):
            #print(f"row index: {r_idx} | row: {row}")
            for c_idx, value in enumerate(row, 1):
                #print(f"column index: {c_idx} | value: {value}")
                cell = self.worksheet.cell(row=r_idx, column=c_idx)
                if type(cell).__name__ is not 'MergedCell':
                    # print(f"Row: {row} | Column: {value}")
                    self.worksheet.cell(row=r_idx, column=c_idx, value=value)

        self.workbook.save(self.filePath)








"""
        inspectionForm = {
            "basicInformation": {
                "Organization": self.inspectionSheetPandas.iloc[self.__basicInformationCellCoords["Organization"][0], self.__basicInformationCellCoords["Organization"][1]].fillna(""),
                "SupplierVendorCode": self.inspectionSheetPandas.iloc[self.__basicInformationCellCoords["Supplier/Vendor Code"][0], self.__basicInformationCellCoords["Supplier/Vendor Code"][1]],
                "PartNumber": self.inspectionSheetPandas.iloc[self.__basicInformationCellCoords["Part Number"][0], self.__basicInformationCellCoords["Part Number"][1]],
                "PartName": self.inspectionSheetPandas.iloc[self.__basicInformationCellCoords["Part Name"][0], self.__basicInformationCellCoords["Part Name"][1]],
                "PurchaseOrder": self.purchaseOrder,
                "DesignRecordChangeLevel": self.inspectionSheetPandas.iloc[self.__basicInformationCellCoords["Design Record Change Level"][0], self.__basicInformationCellCoords["Design Record Change Level"][1]],
                "EngineeringChangeDocuments": self.inspectionSheetPandas.iloc[self.__basicInformationCellCoords["Engineering Change Documents"][0], self.__basicInformationCellCoords["Engineering Change Documents"][1]],
                "RawMaterial": self.inspectionSheetPandas.iloc[self.__basicInformationCellCoords["Raw material"][0], self.__basicInformationCellCoords["Raw material"][1]],
                "Approvedby": self.inspectionSheetPandas.iloc[self.__basicInformationCellCoords["Approved by"][0], self.__basicInformationCellCoords["Approved by"][1]],
                "Supplier": self.inspectionSheetPandas.iloc[self.__basicInformationCellCoords["Supplier"][0], self.__basicInformationCellCoords["Supplier"][1]],
                "Date": self.inspectionSheetPandas.iloc[self.__basicInformationCellCoords["Date"][0], self.__basicInformationCellCoords["Date"][1]]
            },
            "notes": list(self.dataNotesSection.iloc[:, 1].fillna("")),
            "dataToFillByApp": {
                "item": list(self.dataToFillByAppSection.iloc[:, 0]),
                "dimension": list(self.dataToFillByAppSection.iloc[:, 1]), #string
                "units": list(self.dataToFillByAppSection.iloc[:, 2]), #string
                "inferiorLimit": list(self.dataToFillByAppSection.iloc[:, 4]),
                "superiorLimit": list(self.dataToFillByAppSection.iloc[:, 3]),
                "testDate": list(self.dataToFillByAppSection.iloc[:, 5]),
                "quantityTested": list(self.dataToFillByAppSection.iloc[:, 6]), #string
                "measure1": list(self.dataToFillByAppSection.iloc[:, 7]),
                "measure2": list(self.dataToFillByAppSection.iloc[:, 8]),
                "measure3": list(self.dataToFillByAppSection.iloc[:, 9]),
                "measure4": list(self.dataToFillByAppSection.iloc[:, 10]),
                "measure5": list(self.dataToFillByAppSection.iloc[:, 11]),
                "measure6": list(self.dataToFillByAppSection.iloc[:, 12]),
                "ok": list(self.dataToFillByAppSection.iloc[:, 13]),
                "notOk": list(self.dataToFillByAppSection.iloc[:, 14])
            },
            "errorMessage": self.errorMessage
        }
"""