from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import pandas as pd
import numpy as np
import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, Reference, Series

from ..System.SystemService import SystemServices
from Constants.PartNumberConstants import PartNumberConstantsClass
from DeltaHubRestApi.Helpers.AddPartNumber.AddPartNumberHelper import *
from Constants.AddPurchaseOrderConstants import *
from DeltaHubRestApi.Helpers.AddPurchaseOrder.AddPurchaseOrderHelper import AddPurchaseOrderHelpers

class PurchaseOrderExcel:
    def __init__(self):

        self.assignedCells = {
            "A1": "INSPECCION DE CALIDAD",
            "A3": "PO",
            "B3": "",
            "D3": "Fecha de Solicitud",
            "E3": "",
            "A4": "Cantidad",
            "B4": "",
            "D4": "Fecha de Inicio",
            "E4": "",
            "D5": "Fecha de Termino",
            "E5": "",
            "A7": "Piezas",
            "C7": "INSPECCION",
            "A8": "Numero de Parte",
            "B8": "Cant",
            "C8": "Resultado",
            "D8": "Comentarios",
            "G3": "S",
            "G4": "MA",
            "G5": "MI",
            "G6": "OBS",
            "G7": "NE",
            "H3": "X",
            "H4": "X",
            "H5": "X",
            "H6": "X",
            "H7": "X"
        }

    def CreatePurchaseOrderInExcel(self, purchaseOrderForm):
        wb = Workbook()
        ws = wb.active

        cellsToFill = {
            "B3": purchaseOrderForm.purchaseOrderInput,
            "B4": purchaseOrderForm.quantityInput,
            "E3": purchaseOrderForm.requestDateInput,
            "E4": purchaseOrderForm.startDateInput,
            "E5": purchaseOrderForm.endDateInput
        }

        for cell in self.assignedCells:
            if self.assignedCells[cell] is "":
                self.assignedCells[cell] = cellsToFill[cell]

        for cell in self.assignedCells:
            ws[cell] = self.assignedCells[cell]

        return wb

    def AddFileExtensionIfNecessary(self,filePath):
        if ".xlsx" not in filePath and ".XLSX" not in filePath:
            filePath = f"{filePath}.xlsx"
        return filePath

    def ModifyPurchaseOrderInExcel(self, purchaseOrderForm, filePath):
        filePath = self.AddFileExtensionIfNecessary(filePath)
        wb = PurchaseOrderExcel.ReadPurchaseOrderExcel(filePath)
        #wb = Workbook()
        ws = wb.active
        data = ws.values
        cols = next(data)[0:]
        mainDataFrame = PreprocessToMainDataFrame(data, cols)
        mainDataFrameCopy = mainDataFrame.copy()
        rowsSize = mainDataFrameCopy.shape[0]
        columnsSize = mainDataFrameCopy.shape[1]

        print(f"Purchase Order name: {filePath}")
        """
        cellsToFill = {
            [2, 1]: purchaseOrderForm.purchaseOrderInput,
            [3, 1]: purchaseOrderForm.quantityInput,
            [2, 4]: purchaseOrderForm.requestDateInput,
            [3, 4]: purchaseOrderForm.startDateInput,
            [4, 4]: purchaseOrderForm.endDateInput,
        }
        """
        mainDataFrameCopy.iloc[2, 1] = purchaseOrderForm.purchaseOrderInput
        mainDataFrameCopy.iloc[3, 1] = purchaseOrderForm.quantityInput
        mainDataFrameCopy.iloc[2, 4] = purchaseOrderForm.requestDateInput
        mainDataFrameCopy.iloc[3, 4] = purchaseOrderForm.startDateInput
        mainDataFrameCopy.iloc[4, 4] = purchaseOrderForm.endDateInput
        print(mainDataFrameCopy)
        rows = dataframe_to_rows(mainDataFrameCopy, index=False, header=False)
        # print(enumerate(rows, 1))
        # print(list(dataframe_to_rows(mainDataFrame)))
        # for r_idx, row in enumerate(rows, 1):
        #    print(f"row index: {r_idx} | value: {row}")

        for r_idx, row in enumerate(rows, 1):
            # print(f"row index: {r_idx} | row: {row}")
            for c_idx, value in enumerate(row, 1):
                # print(f"column index: {c_idx} | value: {value}")
                cell = ws.cell(row=r_idx, column=c_idx)
                if type(cell).__name__ is not 'MergedCell':
                    # print(f"Row: {row} | Column: {value}")
                    ws.cell(row=r_idx, column=c_idx, value=value)

        """
        for cell in self.assignedCells:
            if self.assignedCells[cell] is "":
                self.assignedCells[cell] = cellsToFill[cell]

        for cell in self.assignedCells:
            ws[cell] = self.assignedCells[cell]
        """
        return wb

    @staticmethod
    def ReadPurchaseOrderExcel(filePath):
        return load_workbook(filename=filePath)

    @staticmethod
    def GetQuantityOfPieces(filePath):
        startIndex = 8
        quantity = 0
        purchaseOrderExcelPandas = pd.read_excel(io=filePath)
        rows = len(purchaseOrderExcelPandas)
        partNumberInfoIndexes = list(np.arange(startIndex, rows))

        listOfQuantities = []

        for cellIndex in partNumberInfoIndexes:
            if purchaseOrderExcelPandas.isnull().iloc[cellIndex, 1]:
                break
            listOfQuantities.append(purchaseOrderExcelPandas.iloc[cellIndex, 1])

        quantity = sum(listOfQuantities)
        return quantity


    @staticmethod
    def GetBasicDataFromPurchaseOrder(filePath, purchaseOrderForm):
        basicInformationCells = {
            "PurchaseOrder": "B3",
            "Quantity": "B4",
            "RequestDate": "E3",
            "StartDate": "E4",
            "EndDate": "E5"
        }
        wb = PurchaseOrderExcel.ReadPurchaseOrderExcel(filePath)
        sheet = wb["Sheet1"]

        purchaseOrderForm.purchaseOrderInput = sheet[basicInformationCells["PurchaseOrder"]].value
        purchaseOrderForm.requestDateInput = PurchaseOrderExcel.ModifyIfValueStringOrDateTime(sheet[basicInformationCells["RequestDate"]].value)
        purchaseOrderForm.startDateInput = PurchaseOrderExcel.ModifyIfValueStringOrDateTime(sheet[basicInformationCells["StartDate"]].value)
        purchaseOrderForm.endDateInput = PurchaseOrderExcel.ModifyIfValueStringOrDateTime(sheet[basicInformationCells["EndDate"]].value)

        purchaseOrderForm.quantityInput = PurchaseOrderExcel.GetQuantityOfPieces(filePath) #Agregar otra celda para TotalCantidad

        return purchaseOrderForm

    @staticmethod
    def ModifyIfValueStringOrDateTime(value):
        # %m/%d/%y
        stringSample = ""
        datetimeSample = datetime.datetime.now()
        if type(value) is type(datetimeSample):
            #value = str(value).replace(" 00:00:00", "")
            #value = str(value).replace("-","/")
            value = str( datetime.datetime.strptime(str(value), "%Y-%m-%d %H:%M:%S").date() ) #convert into date then that date into string

            print(value)

        """
        elif type(value) is type(stringSample):
            #convert from %d/%m/%y ===> %Y-%m-%d
            #comes in string
            dateTime = datetime.datetime.strptime(str(value), "%d-%m-%y")
            #newDate = date

            print(value)
            print(type(value))
            
        """

        return value

class PurchaseOrderManagement:

    def __init__(self, fileName):
        self.writableCells = {
            "H3": "", #Data for S
            "H4": "", #Data for MA
            "H5": "", #Data for MI
            "H6": "", #Data for OBS
            "H7": "", #Data for NE
        }
        self.fileName = fileName
        self.filePath = None
        self.errorMessage = None
        self.workbook = None
        self.worksheet = None
        self.mainDataFrame = None
        self.mainDataFrameCopy = None

        self.__rowWhereDataStarts = 7
        self.writableArea = None
        self.__rowsSize = None
        self.__columnsSize = None

        self.chart = PieChart()
        self.columnForLabelsChart = 7
        self.columnToGetDataForChart = 8
        self.minRowToGetDataForChart = 3
        self.maxRowToGetDataForChart = 7

        self.evaluationCriteria = ["S (SATISFACTORY)",
                              "MA (MAYOR CORRECTIVE)",
                              "MI (MINOR CORRECTIVE)",
                              "OBS (OBSERVATION)",
                              "NE (NOT EVALUATED)"]


    def AddFileExtensionIfNecessary(self):
        if ".xlsx" not in self.fileName and ".XLSX" not in self.fileName:
            self.filePath = f"{self.filePath}.xlsx"
            self.fileName = f"{self.fileName}.xlsx"

    def ReadInspectionSheetExcel(self):
        systemServices = SystemServices(PurchaseOrderServerSideConstants.purchaseOrderFolderParent) #Buils only until path to dekstop

        self.filePath = systemServices.BuildPathForFile(self.fileName) #Build the path until fileName
        self.AddFileExtensionIfNecessary()
        print(self.filePath)

        if not systemServices.CheckIfFilePathExists(self.filePath):
            self.errorMessage = "File does not exist"
            return
        print(f"File Path: {self.filePath}")
        self.workbook = load_workbook(filename=self.filePath)
        self.worksheet = self.workbook.active

        data = self.worksheet.values
        cols = next(data)[0:]
        self.mainDataFrame = PreprocessToMainDataFrame(data, cols)
        self.mainDataFrameCopy = self.mainDataFrame.copy()
        self.__rowsSize = self.mainDataFrameCopy.shape[0]
        self.__columnsSize = self.mainDataFrameCopy.shape[1]


    def UpdateInspectionSheet(self, postedDict, purchaseOrder, inspection):
        #print(self.mainDataFrameCopy)

        self.WriteToDataFrame(postedDict, purchaseOrder, inspection)
        self.WriteToPurchaseOrder()
        self.GraphResults()


    def WriteToDataFrame(self, postedDict, purchaseOrder, inspection):
        print(f"inspection quantity: {inspection.Quantity}")
        partNumberToModify = [[postedDict["BasicInformation"]["PartNumber"], inspection.Quantity, postedDict["BasicInformation"]["Resultado"], postedDict["BasicInformation"]["Comentarios"], None, None, None, None]]
        partNumberSeries = pd.DataFrame(partNumberToModify)
        partNumberText = postedDict["BasicInformation"]["PartNumber"]
        print(f"main dataframe copy: \n{self.mainDataFrameCopy}")
        partNumberExists, rowToUpdate = self.LookForPartNumberInDataFrame(postedDict["BasicInformation"]["PartNumber"])
        if not partNumberExists:
            #if part number does not exist, then add it
            self.mainDataFrameCopy = self.mainDataFrameCopy.append(partNumberSeries, ignore_index=True)
        else:
            #if partnumber already exists, then modify it with location already given
            print(partNumberSeries)
            partNumberToModify = [postedDict["BasicInformation"]["PartNumber"],
                                  purchaseOrder.Quantity,
                                  postedDict["BasicInformation"]["Resultado"],
                                  postedDict["BasicInformation"]["Comentarios"],
                                  None,
                                  None,
                                  None,
                                  None]
            self.mainDataFrameCopy.iloc[rowToUpdate, :] = partNumberToModify
        print(f"main datame copy after adding or modify: \n{self.mainDataFrameCopy}")

        """
        jsonPostedDict["BasicInformation"]["Resultado"]
        inspection.Comments = jsonPostedDict["BasicInformation"]["Comentarios"]
        inspection.ResponsableName = jsonPostedDict["BasicInformation"]["InspectedBy"]
        """

    def LookForPartNumberInDataFrame(self, partNumberToLook):
        #will return if True if there is a part number
        #will return the location in the dataframe where it is
        partNumberExists = False
        rowWhereToUpdate = 7
        for row in self.mainDataFrameCopy.iloc[7:, 0]:
            #print(f"Row: {row == partNumberToLook}")
            if row == str(partNumberToLook):
                partNumberExists = True
                #print(rowWhereToUpdate)
                break
            rowWhereToUpdate += 1
        return partNumberExists, rowWhereToUpdate


    def WriteToPurchaseOrder(self):
        #print(self.mainDataFrameCopy)
        rows = dataframe_to_rows(self.mainDataFrameCopy, index=False, header=False)
        # print(enumerate(rows, 1))
        # print(list(dataframe_to_rows(mainDataFrame)))
        # for r_idx, row in enumerate(rows, 1):
        #    print(f"row index: {r_idx} | value: {row}")

        for r_idx, row in enumerate(rows, 1):
            # print(f"row index: {r_idx} | row: {row}")
            for c_idx, value in enumerate(row, 1):
                # print(f"column index: {c_idx} | value: {value}")
                cell = self.worksheet.cell(row=r_idx, column=c_idx)
                if type(cell).__name__ is not 'MergedCell':
                    # print(f"Row: {row} | Column: {value}")
                    self.worksheet.cell(row=r_idx, column=c_idx, value=value)

        self.workbook.save(self.filePath)

    def GraphResults(self):
        self.ReadInspectionSheetExcel()
        #Get the sum of the elements there are in part number
        #print(f"rows for graph: \n{self.mainDataFrameCopy.iloc[:, 2]}")
        lengthOfResultColumn = len(self.mainDataFrameCopy.iloc[:, 2])
        #print(lengthOfResultColumn)

        #8 is the number of rows which the df would have if there was no data of inspections
        if lengthOfResultColumn <= 8:
            #if there is no data then do not graph, just mark with and X
            self.mainDataFrameCopy[2, 7] = "X"
            self.mainDataFrameCopy[3, 7] = "X"
            self.mainDataFrameCopy[4, 7] = "X"
            self.mainDataFrameCopy[5, 7] = "X"
            self.mainDataFrameCopy[6, 7] = "X"

        else:
            seriesForGraph = AddPurchaseOrderHelpers.GetSeriesForGraph(list(self.mainDataFrameCopy.iloc[8:, 2]), self.evaluationCriteria)

            print(seriesForGraph )
            self.mainDataFrameCopy.iloc[2:7, 6:] = seriesForGraph
            self.WriteToPurchaseOrder()

            self.ReadInspectionSheetExcel()

            labels = Reference(self.worksheet, min_col=7, min_row=3, max_row=7)
            data = Reference(self.worksheet, min_col=8, min_row=2, max_row=7)

            self.chart.add_data(data, titles_from_data=True)

            # set labels in the chart object
            self.chart.set_categories(labels)

            # set the title of the chart
            self.chart.title = "--BY PART NUMBER VISUALIZATION--"

            # add chart to the sheet
            # the top-left corner of a chart
            # is anchored to cell E2 .
            self.worksheet.add_chart(self.chart, "J3")

            self.workbook.save(self.filePath)


    def DeletePartNumberFromPurchaseOrderDoc(self, purchaseOrder, partNumber):
        self.ReadInspectionSheetExcel()
        print(self.filePath)
        print(self.mainDataFrameCopy)

        if ".xlsx" in partNumber:
            partNumber = partNumber.replace(".xlsx", "")

        partNumberExists, rowLocation = self.LookForPartNumberInDataFrame(partNumber)
        if partNumberExists:
            rowToDelete = list(self.mainDataFrameCopy.iloc[rowLocation, :])
            self.mainDataFrameCopy = self.mainDataFrameCopy.drop(labels=rowLocation, axis='index', inplace=False)
            noneArray = np.empty([100, self.mainDataFrameCopy.shape[1]])
            noneArray[:] = None
            noneDataFrame = pd.DataFrame(noneArray)

            self.mainDataFrameCopy = self.mainDataFrameCopy.append(noneDataFrame, ignore_index=True)
            print(rowToDelete)
            self.WriteToPurchaseOrder()

        self.GraphResults()










